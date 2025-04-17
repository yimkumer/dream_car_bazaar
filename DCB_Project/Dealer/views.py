from django.shortcuts import render, HttpResponseRedirect
from ProductManagement.models import *
from UserManagement.models import *
from Customer.models import *
from django.db.models import Exists, F, Subquery, OuterRef, Q
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.http import require_POST
# Create your views here.
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.paginator import Paginator
import xlwt
from django.shortcuts import get_object_or_404


# from django.views import View
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os


from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from django.http import HttpResponse

from django.http import JsonResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from itertools import chain
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import json
from UserManagement.views import *

def share_this(request,car_id):
    # pass
    return render(request, "Dealer/product/car-details-share.html")
    # records = get_object_or_404(CarDetails,id=car_id)
    # record =  CarDetails.objects.get(id=car_id)
    # username = request.user.get_full_name() if request.user.is_authenticated else "Guest"
    #     context = {
    #     'records': record,
    #     'username': username
    # }
    # return render(request, "Dealer/product/invoice.html", context)
    

def upload_pdf(request):
    if request.method == 'POST' and request.FILES['pdf']:
        # Get the uploaded PDF file from the request
        pdf_file = request.FILES['pdf']
        
        # Save the PDF file to your server (or cloud storage)
        file_path = default_storage.save(f"pdfs/{pdf_file.name}", ContentFile(pdf_file.read()))
        file_url = default_storage.url(file_path)

        # Return the URL of the uploaded PDF to the frontend for sharing
        return JsonResponse({'success': True, 'share_url': file_url})

    return JsonResponse({'success': False, 'error': 'No PDF uploaded'})



# def upload_pdf(request):
#     if 'credentials' not in request.session:
#         return redirect('google_drive_auth')

#     credentials = google.oauth2.credentials.Credentials(
#         **request.session['credentials'])

#     drive_service = googleapiclient.discovery.build('drive', 'v3', credentials=credentials)

#     # Step 1: Generate the PDF file
#     file_metadata = {'name': 'shared_page.pdf'}
#     pdf_file = request.FILES['pdf']  # PDF from frontend

#     # Create a temporary file for uploading
#     pdf_path = os.path.join(settings.MEDIA_ROOT, 'pdfs', 'shared_page.pdf')
#     with open(pdf_path, 'wb') as temp_file:
#         temp_file.write(pdf_file.read())

#     media = MediaFileUpload(pdf_path, mimetype='application/pdf')

#     file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()

#     # Step 2: Return the shareable Google Drive link
#     file_id = file.get('id')
#     drive_link = f"https://drive.google.com/file/d/{file_id}/view"

#     # Step 3: Generate the full local URL for the PDF to show in the browser
#     full_pdf_url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, 'pdfs/shared_page.pdf'))

#     return JsonResponse({
#         'success': True,
#         'drive_link': drive_link,
#         'pdf_url': full_pdf_url
#     })

# from django.template.loader import render_to_string
# from django.shortcuts import render


# def invoice_view(request):
#     car_ids = request.GET.getlist('car_ids')  # Adjust to your needs (GET or POST)
#     user = request.user

#     # Fetch car details for selected car IDs and the current user
#     car_details_records = CarDetails.objects.filter(id__in=car_ids, created_by=user)

#     # Prepare a list to store rendered templates for each car record
#     rendered_templates = []

#     # Loop over each car record and render its own template
#     for car in car_details_records:
#         car_images = CarImage.objects.filter(car_detail=car, is_active=True)

#         # Render the individual car template as a string
#         car_template = render_to_string('Dealer/product/car_invoice.html', {
#             'record': car,
#             'car_images': car_images
#         })

#         # Append the rendered template to the list
#         rendered_templates.append(car_template)

#     # Combine all the rendered templates into one
#     combined_invoices = ''.join(rendered_templates)

#     # Optionally, render the final template that includes all the combined content
#     return render(request, 'Dealer/product/invoice-combined.html', {
#         'combined_invoices': combined_invoices
#     })




def invoice_view(request):
    selected_car_ids = request.POST.getlist('car_ids')
    if selected_car_ids:
        records = CarDetails.objects.filter(id__in=selected_car_ids,created_by= request.user)
    else:
        records = CarDetails.objects.filter(created_by= request.user)
        
    username = request.user.get_full_name() if request.user.is_authenticated else "Guest"
    contact = request.user.phone
    address = request.user.address
    d = {}
    car_images = []
    for record in records:
            images = CarImage.objects.filter(car_detail = record)
            d[record] = images
            
    # Pass the record and username to the template
    context = {
        'records': records,
        'username': username,
        'contact': contact,
        'address':address,
        'd': d,
    }
    
    return render(request,"Dealer/product/car-details-share.html", context)
    # return render(request, "Dealer/product/invoice.html", context)


def download_selected_pdf(request):
    # Get selected car_ids from the POST request
    selected_car_ids = request.POST.getlist('car_ids')

    if not selected_car_ids:
        return HttpResponse("No records selected.", status=400)

    # Fetch only the selected car records
    selected_cars = CarDetails.objects.filter(id__in=selected_car_ids,created_by= request.user)

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="selected_cars.pdf"'

    # Generate the PDF using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter  # Get the page size

    # Add table headers (5 columns now, including Date)
    p.setFont("Helvetica-Bold", 10)  # Set font size
    # Add table headers
    p.drawString(30,750,"SNo.")
    p.drawString(100, 750, "Car Name")
    p.drawString(290,750, "Registration No.")
    p.drawString(420, 750, "Date")
    p.drawString(510,750,"Car Price(INR)")
    p.setFont("Helvetica", 9)

    # Add table rows
    y = 730
    count = 1
    for car in selected_cars:
        if car.status == "deleted":
            continue
        else:
            p.drawString(30,y,str(count))
            p.drawString(100, y, str(car.variant.model.brand.name + " " +car.variant.model.name+ " " + car.variant.name))
            p.drawString(290, y, str(car.registration_no))
            created_at_str = car.created_at.strftime('%d-%m-%Y')
            p.drawString(420, y, created_at_str)
            p.drawString(510, y, str(car.demand_price))
            y -= 20
            count += 1
    # Finish up the PDF
    p.showPage()
    p.save()

    return response



def download_pdf(request):
    page_number = request.GET.get('page', 1)
    paginator = Paginator(CarDetails.objects.filter(created_by= request.user), 10)  # 10 items per page
    cars = paginator.get_page(page_number)

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="car_list.pdf"'

    # Generate the PDF using ReportLab
    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 10)  # Set font size
    # Add table headers
    p.drawString(30,750,"SNo.")
    p.drawString(100, 750, "Car Name")
    p.drawString(290,750, "Registration No.")
    p.drawString(420, 750, "Date")
    p.drawString(510,750,"Car Price(INR)")
    p.setFont("Helvetica", 9)

    # Add table rows
    y = 730
    count = 1
    for car in cars:
        if car.status == "deleted":
            continue
        else:
            p.drawString(30,y,str(count))
            p.drawString(100, y, str(car.variant.model.brand.name + " " +car.variant.model.name+ " " + car.variant.name))
            p.drawString(290, y, str(car.registration_no))
            created_at_str = car.created_at.strftime('%d-%m-%Y')
            p.drawString(420, y, created_at_str)
            p.drawString(510, y, str(car.demand_price))
            y -= 20
            count += 1
    # Finish up the PDF
    p.showPage()
    p.save()

    return response




def download_selected_excel(request):
    # Get selected car_ids from the POST request
    selected_car_ids = request.POST.getlist('car_ids')

    if not selected_car_ids:
        return HttpResponse("No records selected.", status=400)

    # Fetch only the selected car records
    selected_cars = CarDetails.objects.filter(id__in=selected_car_ids,created_by= request.user)

    # Create the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="selected_cars.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Cars')

    # Add headers to Excel file
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['SNO. ', 'Car Name', 'Registration No.', 'Date' ,'Car Price(INR)']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Add data to Excel file
    font_style = xlwt.XFStyle()
    count = 1
    for car in selected_cars:
        if car.status == "deleted":
            continue
            
        else:
            row_num += 1
            name = str(car.variant.model.brand.name + " " +car.variant.model.name+ " " + car.variant.name)
            created_at_str = car.created_at.strftime('%d-%m-%Y')
            ws.write(row_num, 0, count, font_style)
            ws.write(row_num, 1, name, font_style)
            ws.write(row_num, 2, car.registration_no, font_style)
            ws.write(row_num, 3, created_at_str, font_style)
            ws.write(row_num, 4, car.demand_price, font_style)
            count += 1

    wb.save(response)
    return response


def download_excel(request):
    
    page_number = request.GET.get('page', 1)  # Default to page 1 if not specified
    # Add data from the database
    cars = CarDetails.objects.filter(created_by= request.user)  # Fetch your records

    # Paginate the records
    paginator = Paginator(cars, 10)  # Show 10 cars per page
    page_obj = paginator.get_page(page_number)
    # Create Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="records.xlsx"'

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Records'

    # Add headers
    # ws.append(['SNO. ', 'Car Name', 'Registration No.', 'Date' ,'Car Price(INR)'])  # Customize with your headers
    columns = ['SNO. ', 'Car Name', 'Registration No.', 'Date' ,'Car Price(INR)']

    count = 1
    
    # Create bold font style for headings
    bold_font = Font(bold=True)

    # Write column headings
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font

    # Write data
    for row_num, car in enumerate(page_obj.object_list, 2):
        if car.status == "deleted":
            continue
        else:
            name = str(car.variant.model.brand.name + " " +car.variant.model.name+ " " + car.variant.name)
            created_at_str = car.created_at.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=1, value=count)
            ws.cell(row=row_num, column=2, value=name if name is not None else '')
            ws.cell(row=row_num, column=3, value=car.registration_no if car.registration_no is not None else '')
            ws.cell(row=row_num, column=4, value=created_at_str if created_at_str is not None else '')
            ws.cell(row=row_num, column=5, value=car.demand_price if car.demand_price is not None else '')
            count += 1

    # Save the workbook to the response
    wb.save(response)
    return response





def add_car(request):
    user = request.user
    # all_car_details = CarDetails.objects.filter(created_by=user)
    car_brand = CarBrands.objects.filter(is_active=True)
    car_model = CarModel.objects.filter(is_active=True)
    car_variant = CarVariant.objects.filter(is_active=True)
        
    return render(request, "Dealer/product/add_car.html", {'car_brands': car_brand,
                                                                 'car_models': car_model,
                                                                 'car_variants': car_variant})


def marksold(request,id):
    if request.method == 'POST':
        car = get_object_or_404(CarDetails,id=id)
        car.status = "sold"
        car.save()
        messages.info(request,'Marked as Sold !')
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)

    return HttpResponse(status=405)

def markdelete(request,id):
    if request.method == 'POST':
        today = timezone.now().date()
        car = get_object_or_404(CarDetails, id=id)
        car.previous_status = car.status
        car.status = "deleted"
        car.deletion_date = today
        car.save()
        messages.info(request,"Deleted Successfully!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)
    return HttpResponse(status=405)

def repost_ad(request,id):
    if request.method == 'POST':
        car = get_object_or_404(CarDetails, id=id)
        # Update the car status to 'live' or appropriate status
        if car.status == "deleted":
            car.status = car.previous_status
            car.previous_status = None
            car.save()
            
        if car.previous_status == None:
            pass
        messages.info(request, "Reposted Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)
          
    return HttpResponse(status=405)  # Method not allowed if not POST

def expired_ad(request):
    pass

@require_POST
def delete_ad(request, id):
    try:
        item = get_object_or_404(CarDetails, id = id)
        item.delete()
        messages.info(request, "Post Permanent Delete Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)        
        # return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'fail', 'error': str(e)}, status=400)
        


def update_car_detail(request, id):
    car_detail = get_object_or_404(CarDetails, id=id)

    if request.method == 'POST':
        # Update the fields with the form data
        car_detail.insurance_type = request.POST.get('insurance_type')
        car_detail.insurance_by = request.POST.get('insurance_by')
        car_detail.fitness_validity = request.POST.get('fitness_validity')
        car_detail.pollution_validity = request.POST.get('pollution_validity')
        car_detail.registration_place = request.POST.get('registration_place')
        car_detail.registration_no = request.POST.get('registration_no')
        car_detail.last_service_date = request.POST.get('last_service')
        car_detail.vehicle_location = request.POST.get('vehicle_location')
        car_detail.save()  # Save changes to the database
        messages.success(request, 'Car details updated successfully!')
            # return redirect(request) 
    else:
        messages.error(request,'Invalid form Data, Try again!')
    
    return render(request, "Dealer/product/car_details.html", {'car_detail': car_detail})
    
    
def update_interior(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':
        car_detail.power_steering = request.POST.get('power_steering')
        car_detail.cruise_control = request.POST.get('cruise_control')
        car_detail.navigation_system = request.POST.get('navigation_system')
        car_detail.adjustable_steering = request.POST.get('adjustable_steering')
        car_detail.steering_control = request.POST.get('steering_control')
        car_detail.air_conditioning = request.POST.get('air_conditioning')
        car_detail.power_window = request.POST.get('power_window')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
    else:
        messages.error(request,'Invalid form Data, Try again!')

    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})
    

def update_exterior(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':
        car_detail.alloy_wheel = request.POST.get('alloy_wheel')
        car_detail.sun_roof = request.POST.get('sun_roof')
        car_detail.adjustable_orvm = request.POST.get('adjustable_orvm')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
    else:
        messages.error(request,'Invalid form Data, Try again!')

    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})
    
def update_infotainment(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':
        car_detail.bluetooth = request.POST.get('bluetooth')
        car_detail.am_fm_radio = request.POST.get('am_fm_radio')
        car_detail.usb_compatibility = request.POST.get('usb_compatibility')
        car_detail.aux_compatibility = request.POST.get('aux_compatibility')
        car_detail.wireless_charger = request.POST.get('wireless_charger')
        car_detail.android_car_play = request.POST.get('android_car_play')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
    else:
        messages.error(request,'Invalid form Data, Try again!')

    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})
    
def update_safety(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':   
        car_detail.abs = request.POST.get('abs')
        car_detail.ebd = request.POST.get('ebd')
        car_detail.anti_theft_device = request.POST.get('anti_theft_device')
        car_detail.rear_parking_camera = request.POST.get('rear_parking_camera')
        car_detail.rear_parking_sensor = request.POST.get('rear_parking_sensor')
        car_detail.front_parking_camera = request.POST.get('front_parking_camera')
        car_detail.lock_system = request.POST.get('lock_system')
        car_detail.total_air_bag = request.POST.get('total_air_bag')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
    else:
        messages.error(request,'Invalid form Data, Try again!')

    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})
    
    
    
    
def update_vehicle(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':    
        car_detail.battery_status = request.POST.get('battery_status')
        car_detail.tyre_condition = request.POST.get('tyre_condition')
        car_detail.vehicle_warranty = request.POST.get('vehicle_warranty')
        car_detail.vehicle_warranty_date = request.POST.get('vehicle_warranty_date')
        car_detail.accidental = request.POST.get('accidental')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
    else:
        messages.error(request,'Invalid form Data, Try again!')

    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})


def update_additional(request,id):
    car_detail = get_object_or_404(CarDetails, id=id)
    if request.method == 'POST':
        car_detail.finance = request.POST.get('finance')
        car_detail.exchange = request.POST.get('exchange')
        car_detail.description = request.POST.get('description')
        car_detail.extended_warranty = request.POST.get('extended_warranty')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
        # return redirect(request)
    else:
        messages.error(request, 'Invalid form Data, Try again!')

        
    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})
    
    
def update_top(request,id):
    car_detail = get_object_or_404(CarDetails,id = id)
    if request.method == 'POST':
        car_detail.color = request.POST.get('color')
        car_detail.ownership = request.POST.get('ownership')
        car_detail.fuel_type = request.POST.get('fuel_type')
        car_detail.engine_capacity = request.POST.get('engine_capacity') 
        car_detail.demand_price = request.POST.get('demand_price')
        car_detail.transmission = request.POST.get('transmission')
        car_detail.total_km_driven = request.POST.get('total_km_driven')
        car_detail.brand_name = request.POST.get('brand_name')
        car_detail.model_name = request.POST.get('model_name')
        car_detail.variant_name = request.POST.get('variant_name')
        car_detail.save()
        messages.success(request,'Car details updated successfully!')
        # return redirect(request)
    else:
        messages.error(request, 'Invalid form Data, Try again!')

        
    return render(request,"Dealer/product/car_details.html",{'car_detail':car_detail})        

def car_list(request):
    user = request.user
    all_cars = CarDetails.objects.filter(created_by=user, review='approved').exclude(status='deleted').order_by('-updated_at')
    # made_live = DraftCarDetails.objects.filter(is_draft=False)
    all_car_details = list(all_cars)
    cars_count = len(all_car_details)
    # all_car_details = CarDetails.objects.exclude(status = 'deleted')
    # car_id = request.POST.get('car_detail.id')
    live = CarDetails.objects.filter(created_by=user,status="live",review='approved').order_by('-updated_at')
    live_count = len(live)
    expired = CarDetails.objects.filter(created_by=user,status="expired").order_by('-updated_at')
    expired_count = len(expired)
    sold = CarDetails.objects.filter(created_by=user,status="sold").order_by('-updated_at')
    sold_count = len(sold)
    spams = SpamReport.objects.all()
    deleted_cars = CarDetails.objects.filter(created_by=user,status="deleted").order_by('-created_at')
    # drafts = DraftCarDetails.objects.filter(created_by=user,is_draft=True).exclude(status='deleted').exclude(review='approved').exclude(review='rejected').order_by('-updated_at')
    drafts = DraftCarDetails.objects.all()
    processing = CarDetails.objects.filter(created_by=user,status='processing').order_by('-updated_at')
    alldrafts = list(chain(drafts,processing))
    deleted_drafts = DraftCarDetails.objects.filter(created_by=user,status='deleted').order_by('-created_at')
    deleted = list(chain(deleted_cars,deleted_drafts))
    
    car_brand = CarBrands.objects.filter(is_active=True)
    car_model = CarModel.objects.filter(is_active=True)
    car_variant = CarVariant.objects.filter(is_active=True)

    for item in deleted:
        item.status = "deleted"
        item.save()
    deleted_count = len(deleted)
    draft_count = len(alldrafts)
    
    # for spam in spams:
    #     temp = spam.draft_id
    #     res = f"{temp}"
    #     item = CarDetails.objects.get(id=res)
    #     item.remarks = spam.remarks
    #     item.user = spam.user.get_full_name()
    #     if item:
    #         rejected.append(item)
    #     else:
    #         continue
    # gallery = []
    rejected = CarDetails.objects.filter(created_by=user, review='rejected').exclude(status='deleted').order_by('-updated_at')
    # for draft in drafts:
    #     img = CarImage.objects.filter(car_detail= draft)
    #     if img:
    #         # draft.images = img
    #         gallery.append(img)
    
    return render(request, "Dealer/product/car_list.html", {'all_car_detail': all_car_details,'draft_count':draft_count,'drafts': alldrafts,'sold':sold,'deleted':deleted,'live':live,'expired':expired,
    'rejected':rejected,'cars_count':cars_count,'live_count':live_count,'expired_count':expired_count,'sold_count':sold_count,'deleted_count':deleted_count,
    'car_brands': car_brand,'car_models': car_model,'car_variants': car_variant})



def car_detail(request, id):
    if request.user.is_authenticated:
        user = request.user
        car_details = CarDetails.objects.get(id=id)
        car_images = CarImage.objects.filter(car_detail= car_details)
        print('car_details = ', car_details.variant)
        car_brand = CarBrands.objects.filter(is_active=True)
        car_model = CarModel.objects.filter(is_active=True)
        car_variant = CarVariant.objects.filter(is_active=True)
        return render(request, "Dealer/product/car_details.html", {'car_detail': car_details,'car_images':car_images,'car_brands': car_brand,'car_models': car_model,'car_variants': car_variant})

@csrf_exempt
def update_duration(request):
    group_list = request.user.groups.values_list('name', flat=True)
    
    if request.method == 'POST' and request.user.is_authenticated:
        data = json.loads(request.body)
        time_spent = data.get('time_spent', 0)
        car = data.get('car','')
        post = get_object_or_404(CarDetails, id=car)
        # Retrieve or create the Lead object for the current user
        lead, created = Lead.objects.get_or_create(user=request.user, viewed_car=post)
        # Update the count of visit
        
        
        if not created:
            # If the object already exists, modify the fields you want to update
            lead.viewed_car = post
            lead.status = 'cold'
            lead.visit_count += 1
            lead.user_type = group_list[0]
        else:
            lead.visit_count = 1
        
        lead.page_stay_duration += time_spent
        lead.user_type = group_list[0]
        
        lead.save()

        return JsonResponse({'message': 'Duration updated'})
    return JsonResponse({'error': 'Invalid request'}, status=400)
    
def update_lead_status(request,car):
    post = get_object_or_404(CarDetails, id=car)
    owner = post.created_by
    group_list = request.user.groups.values_list('name', flat=True)
    today = timezone.now().date()
    # lead, created = Lead.objects.get_or_create(user=request.user, viewed_car=post)
    if request.method == 'POST':
        action = request.POST.get('action')
        day = request.POST.get('std_day')
        
        
        if action == 'contact_seller':
            lead, created = Lead.objects.get_or_create(user=request.user, viewed_car=post, defaults={'status': 'warm', 'user_type': group_list[0]})
            if not created:
                lead.status = 'warm'
                lead.visit_count += 1
                lead.user_type = group_list[0]
            else:
                lead.visit_count = 1
            lead.save()
            messages.success(request, 'Our Team may contact you soon !')

        elif action == 'schedule_test_drive':
            lead, created = Lead.objects.get_or_create(user=request.user, viewed_car=post, defaults={'visit_time': day ,'status':'hot','user_type': group_list[0]})
            notification, notify = Notification.objects.get_or_create(user=request.user, car=post)
            # Only set the message once, no matter if the notification was created or fetched
            if (day-today).days <= 2:
                msg = f" Hurry {(day-today).days} Days left! {lead.user.get_full_name()} scheduled a test drive on {day} for {lead.viewed_car.variant.model.brand.name}{lead.viewed_car.variant.model.name}{lead.viewed_car.variant.name}"
                notification.message = msg
            else:
                notification.message = f"{lead.user.get_full_name()} scheduled a test drive on {day} for {lead.viewed_car.variant.model.brand.name}{lead.viewed_car.variant.model.name}{lead.viewed_car.variant.name}"
            
            notification.save()

            if not created:
                lead.status = 'hot'
                lead.visit_time = day
                lead.visit_count += 1
                lead.user_type = group_list[0]
                # diff = (lead.visit_time - today).days
                # if diff <= 2:
                #     msg = f" Hurry {diff} Days left! {lead.user.get_full_name()} scheduled a test drive on {day} for {lead.viewed_car.variant.model.brand.name}{lead.viewed_car.variant.model.name}{lead.viewed_car.variant.name}"
                #     alert, flag = Notification.objects.get_or_create(user=request.user, car=post)
                #     alert.message = msg
                #     alert.save()
            else:
                lead.visit_count = 1
            lead.save()
            messages.success(request, 'Our Team will contact you soon !!!')
        
        return  HttpResponseRedirect(request.path)
            
    return HttpResponseRedirect('/')
  

def follow_up_view(request):
    cold_leads = Lead.objects.filter(status='cold', viewed_car__created_by=request.user)
    warm_leads = Lead.objects.filter(status='warm', viewed_car__created_by=request.user)
    hot_leads = Lead.objects.filter(status='hot', viewed_car__created_by=request.user)
    cars = CarDetails.objects.filter(created_by=request.user)

    # Create a list of cars with concatenated brand, model, and variant
    car_list = []
    for car in cars:
        full_name = f"{car.variant.model.brand.name} {car.variant.model.name} {car.variant.name}"
        car_list.append({
            'id': car.id,  # You can also pass the car ID or any other field
            'full_name': full_name
        })

    selected_car = request.GET.get('car_name')

    if selected_car and selected_car != 'all':
        selected_car_parts = selected_car.split(' ', 2)
        if len(selected_car_parts) == 3:
            brand_name, model_name, variant_name = selected_car_parts

            # Filter cars based on the selected brand, model, and variant
            cars = CarDetails.objects.filter(
                Q(variant__model__brand__name=brand_name) &
                Q(variant__model__name=model_name) &
                Q(variant__name=variant_name)
            )

            # Filter leads based on the selected car (by matching interest with the selected car)
            cold_leads = cold_leads.filter(
                Q(viewed_car__variant__model__brand__name=brand_name) &
                Q(viewed_car__variant__model__name=model_name) &
                Q(viewed_car__variant__name=variant_name)
            )
            warm_leads = warm_leads.filter(
                Q(viewed_car__variant__model__brand__name=brand_name) &
                Q(viewed_car__variant__model__name=model_name) &
                Q(viewed_car__variant__name=variant_name)
            )
            hot_leads = hot_leads.filter(
                Q(viewed_car__variant__model__brand__name=brand_name) &
                Q(viewed_car__variant__model__name=model_name) &
                Q(viewed_car__variant__name=variant_name)
            )
    
    # If no car is selected or 'all' is selected, show all leads.
    else:
        cars = CarDetails.objects.filter(created_by=request.user)
        cold_leads = Lead.objects.filter(status='cold', viewed_car__created_by=request.user)
        warm_leads = Lead.objects.filter(status='warm', viewed_car__created_by=request.user)
        hot_leads = Lead.objects.filter(status='hot', viewed_car__created_by=request.user)
    

    context = {
        'cold_leads': cold_leads,
        'warm_leads': warm_leads,
        'hot_leads': hot_leads,
        'cars': car_list,  # Pass the modified car list to the template
        'selected_car': selected_car,
    }
    return render(request, 'Dealer/followup/followup.html', context)


def mark_as_read(request, notification_id):
    # Fetch the specific notification by ID for the logged-in user
    notification = get_object_or_404(Notification, id=notification_id)

    # Mark the notification as read if it hasn't been read already
    if not notification.is_read:
        notification.is_read = True
        notification.save()
    
    return redirect(request.META.get('HTTP_REFERER', '/'))

def delete_notification(request, notification_id):
    if request.method == 'POST':
        notification = get_object_or_404(Notification, id=notification_id)
        notification.delete()
    return redirect(request.META.get('HTTP_REFERER', '/'))

def custom_404(request, exception):
    return render(request, "Dealer/error/404.html", status=404)

def download_lead_excel(request):
    
    page_number = request.GET.get('page', 1)  # Default to page 1 if not specified
    # Add data from the database
    cars = Lead.objects.filter(viewed_car__created_by= request.user)  # Fetch your records

    # Paginate the records
    paginator = Paginator(cars, 10)  # Show 10 cars per page
    page_obj = paginator.get_page(page_number)
    # Create Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="leads.xlsx"'

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Records'

    # Add headers
    # ws.append(['SNO. ', 'Car Name', 'Registration No.', 'Date' ,'Car Price(INR)'])  # Customize with your headers
    columns = ['SNO. ','User', 'User Type', 'Phone','Email', 'Visited At','Engagement(in Seconds)', 'Status', 'Interest', 'Visit Count']

    count = 1
    
    # Create bold font style for headings
    bold_font = Font(bold=True)

    # Write column headings
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font

    # Write data
    for row_num, lead in enumerate(page_obj.object_list, 2):
            user = lead.user.get_full_name()
            car_name = str(lead.viewed_car.variant.model.brand.name + " " +lead.viewed_car.variant.model.name+ " " + lead.viewed_car.variant.name)
            
            ws.cell(row=row_num, column=1, value=count)
            ws.cell(row=row_num, column=2, value=user if user is not None else '')
            ws.cell(row=row_num, column=3, value=lead.user_type if lead.user_type is not None else '')
            ws.cell(row=row_num, column=4, value=lead.user.phone if lead.user.phone is not None else '')
            ws.cell(row=row_num, column=5, value=lead.user.email if lead.user.email is not None else '')
            ws.cell(row=row_num, column=6, value=str(lead.visit_time) if lead.visit_time is not None else '')
            ws.cell(row=row_num, column=7, value=lead.page_stay_duration if lead.page_stay_duration is not None else '')
            ws.cell(row=row_num, column=8, value=lead.status if lead.status is not None else '')
            ws.cell(row=row_num, column=9, value=car_name if car_name is not None else '')
            ws.cell(row=row_num, column=10, value=lead.visit_count if lead.visit_count is not None else '')            
            count += 1

    # Save the workbook to the response
    wb.save(response)
    return response

def customer_car(request):
    if request.user.is_authenticated:
        user = request.user
        group_list = user.groups.values_list('name', flat=True)
        car_image = CarImage.objects.filter(car_detail=OuterRef('id')).order_by('-created_at').values('image')[:1]
       
        # Filter to show only cars created by the customer
        if 'Customer' in group_list:
            all_car_detail = CarDetails.objects.filter(created_by=user, is_active=True).annotate(car_image=Subquery(car_image)).order_by('-created_at')
        else:
            all_car_detail = CarDetails.objects.filter(is_active=True).exclude(created_by=user).annotate(car_image=Subquery(car_image)).order_by('-created_at')

        print('all_car_detail = ', all_car_detail)

        if 'Dealer' in group_list:
            return render(request, "Dealer/product/customer_car.html", {'all_car_detail': all_car_detail})
        else:
            return HttpResponseRedirect('/profile/')
    else:
        return HttpResponseRedirect('/')








def profile(request):
    user = request.user
    if not user.is_authenticated:
        messages.info(request,"Please login to access the page!")
        return home(request)

    if request.method == 'POST':
        user_name = request.POST.get('dealer_name')
        firm_name = request.POST.get('dealer_firm_name')
        edit = request.POST.get('edit')
        email = request.POST.get('dealer_email')
        address = request.POST.get('dealer_address')
        dp = request.FILES.get('dealer_dp')
        aadhar = request.FILES.get('dealer_aadhar')
        pan = request.FILES.get('dealer_pan')
        gst = request.FILES.get('dealer_gst')
        if edit == 'edit_name':
            CustomUser.objects.filter(id=user.id).update(firm_name=firm_name,first_name=user_name)
        elif edit == 'edit_email':
            CustomUser.objects.filter(id=user.id).update(email=email)
        elif edit == 'edit_address':
            CustomUser.objects.filter(id=user.id).update(address=address)
        elif edit == 'edit_DP' and dp:
            user.user_image = dp  # Assign the uploaded image to the user instance
            user.save()
        elif edit == 'update_card' and aadhar and pan and gst:
            user.aadhar_card = aadhar
            user.pan_card = pan
            user.gst = gst
            user.save()
        elif edit == 'edit_pass':
            current_password = request.POST.get('dealer_old_password')
            new_password = request.POST.get('dealer-new-password')
            confirm_password = request.POST.get('dealer-re-new-password')
            if user.check_password(current_password):
                user.set_password(new_password)
                user.save()
                update_session_auth_hash(request, user)
                
                
        messages.info(request, "Profile Edit Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)
        
    return render(request, "Dealer/profile/profile_page.html")
    

def dealer_dashboard(request):
    user = request.user
    total_car = CarDetails.objects.filter(created_by=user).count()
    # views.py
    notifications = Notification.objects.filter(car__created_by=user ,is_read=False).order_by('-created_at')

    # print('car_details = ', car_details.variant)
    return render(request, "Dealer/dashboard/dashboard.html", {'user': user, 'total_car': total_car, 'notifications':notifications})

def expired_insurances(request):
    today = timezone.now().date()

    insurances = Insurance.objects.filter(created_by= request.user)
    current = []
    expired = []

    for insurance in insurances:
        if insurance.risk_start_date is None:
            continue
        expiry_date = insurance.risk_start_date + relativedelta(months=12) -relativedelta(days=1)
        reminder_date = expiry_date - timedelta(days=5)
        # Append insurance object to the appropriate list
        if expiry_date >= today:
            if insurance.get_status() == 'live':
                current.append((insurance,expiry_date,reminder_date))
        else:
            insurance.status = insurance.get_status()
            expired.append((insurance,expiry_date,reminder_date))
    
    expired.sort(key=lambda x: x[1], reverse = True)
    # Render the template with the sorted warranties
    return expired, current

def sorted_insurances(request):
    # Get the current date
    today = timezone.now().date()

    # Calculate the start and end dates for the filtering
    first_day_of_current_month = today.replace(day=1)
    last_day_of_current_month = (first_day_of_current_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Extend the range by 5 days before and after
    filter_start_date = first_day_of_current_month - timedelta(days=5)
    filter_end_date = last_day_of_current_month + timedelta(days=5)

    # Filter  by expiry date
    filtered_insurances = Insurance.objects.filter(
        risk_start_date__isnull=False,
        created_by = request.user
    )

    # Calculate expiry date for each  and filter based on the calculated range
    renewal = []
    expired = []
    for insurance in filtered_insurances:
        if insurance.risk_start_date is None:
            continue
        expiry_date = insurance.risk_start_date + relativedelta(months=12) - relativedelta(days=1)
        reminder_date = expiry_date - timedelta(days=5)
        if filter_start_date <= expiry_date <= filter_end_date:
            insurance.status = insurance.get_status()
            renewal.append((insurance, expiry_date,reminder_date))
        elif expiry_date < filter_start_date:
            expired.append((insurance, expiry_date,reminder_date))
    # Sort  by expiry date
    renewal.sort(key=lambda x: x[1])
    expired.sort(key=lambda x:x[1])
    
    return renewal

from django.views.decorators.http import require_POST
from django.shortcuts import redirect

@require_POST
def delete_insurance(request, id):
    try:
        insurance = get_object_or_404(Insurance, id = id)
        insurance.delete()
        messages.info(request, "Insurance Deleted Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)        
        # return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'fail', 'error': str(e)}, status=400)
    
    return render(request, "Dealer/insurance/insurance.html")
# import pandas as pd

def insurance(request):
    user = request.user
    car_brand = CarBrands.objects.filter(is_active=True)
    # insurance_list = Insurance.objects.filter(is_active=True)
    insurance_list = Insurance.objects.filter(created_by=user).order_by('-risk_end_date')
    insurance_count = insurance_list.count()
    for insurance in insurance_list:
        insurance.status = insurance.get_status()    
        insurance.save()
    # file_path = os.path.join(settings.BASE_DIR, 'templates', 'Dealer', 'insurance', 'insurer.csv')
    
    # # Check if the file exists at the path
    # if not os.path.exists(file_path):
    #     raise FileNotFoundError(f"File not found at {file_path}")
   
    # df = pd.read_csv(file_path)
    # # print(df.columns)
    # companies = df['Company name'].tolist()
    companies = InsuranceCompany.objects.all()
    insurance_id = request.POST.get('insurance_id')
    if insurance_id:
        # Edit existing warranty
        insurance = get_object_or_404(Insurance, pk=insurance_id)
        for field, value in insurance_dict.items():
            setattr(Insurance, field, value)
        Insurance.save()  # Save the updated warranty to the database   

    if request.method == 'POST' and not insurance_id:
        user_phone_no = request.POST.get('user_phone_no')
        # user_name = request.POST.get('user_name')
        policy_no = request.POST.get('policy_no')
        insured_name = request.POST.get('user_name')
        # insurance_date = request.POST.get('insurance_date')
        type = request.POST.get('type')
        product = request.POST.get('product')
        policy_type = request.POST.get('policy_type')
        risk_start_date = request.POST.get('risk_start_date')
        risk_end_date = request.POST.get('risk_end_date')
        ncb_status = request.POST.get('ncb_status')
        idv = request.POST.get('idv')
        insurer_name = request.POST.get('company')
        car_model = request.POST.get('car_model')
        fuel_type = request.POST.get('fuel_type')
        next_renewal_date = request.POST.get('next_renewal_date') + "-01"
        rto_state = request.POST.get('rto_state')
        city = request.POST.get('city')
        car_model_instance = CarModel.objects.filter(id=car_model)
        total_premium = request.POST.get('total_premium')
        created_by = user

        if car_model_instance.exists():
            insurance_dict = {'user_phone_no': user_phone_no,'insured_name': insured_name , 'policy_no': policy_no,
            'insurer_name':insurer_name ,'city': city, 'type': type, 'product': product,
            'policy_type': policy_type, 'risk_start_date': risk_start_date, 'risk_end_date': risk_end_date,
            'ncb_status': ncb_status, 'idv': idv, 'total_premium': total_premium, 'car_model': car_model_instance.last(),
            'fuel_type': fuel_type, 'next_renewal_date': next_renewal_date, 'rto_state': rto_state, 'created_by': created_by,'is_active': 1}
            # print()
            Insurance.objects.create(**insurance_dict)
            messages.info(request, "Insurance Form Submit Successfully !!")
            referrer = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(referrer)
        

    # car_details = CarDetails.objects.get(id=id)
    # print('car_details = ', car_details.variant)
    insurances_with_expiry = sorted_insurances(request)
    expired, live = expired_insurances(request)  
    live_count = len(live)
    expired_count = len(expired)
    renewal_count = len(insurances_with_expiry)
    return render(request, "Dealer/insurance/insurance.html", {'companies':companies,'car_brand': car_brand,'insurance_list':insurance_list,'renewal_count':renewal_count,'insurance_count':insurance_count,'expired_count':expired_count,
                            'expired':expired,'renewal':insurances_with_expiry, 'current':live ,'live_count':live_count })



from django.utils import timezone
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from django.shortcuts import get_object_or_404,redirect
from django.http import JsonResponse
from ProductManagement.models import Warranty

def expired_warranty(request):
    today = timezone.now().date()
    user = request.user
    # Fetch all warranties and compute their expiry dates
    warranties = Warranty.objects.filter(created_by=user)
    current_warranties = []
    expired_warranties = []

    for warranty in warranties:
        expiry_date = warranty.purchase_date + relativedelta(months=warranty.warranty_period) -relativedelta(days=1)
        reminder_date = expiry_date - timedelta(days=5)
        # Append warranty object to the appropriate list
        if expiry_date >= today:
            if warranty.get_status() == 'live':
                current_warranties.append((warranty,expiry_date,reminder_date))
        else:
            # warranty.status = warranty.get_status()
            # warranty.save()
            expired_warranties.append((warranty,expiry_date,reminder_date))
    
    expired_warranties.sort(key=lambda x: x[1])
    
    # Render the template with the sorted warranties
    return expired_warranties, current_warranties

def sorted_warranty_users(request):
    # Get the current date
    today = timezone.now().date()
    user = request.user
    # Calculate the start and end dates for the filtering
    first_day_of_current_month = today.replace(day=1)
    last_day_of_current_month = (first_day_of_current_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Extend the range by 5 days before and after
    filter_start_date = first_day_of_current_month - timedelta(days=5)
    filter_end_date = last_day_of_current_month + timedelta(days=5)

    # Filter warranties by expiry date
    filtered_warranties = Warranty.objects.filter(
        purchase_date__isnull=False,
        warranty_period__in=[6, 12],
        created_by = user
    )

    # Calculate expiry date for each warranty and filter based on the calculated range
    warranties_in_range = []
    for warranty in filtered_warranties:
        # expiry_date = warranty.purchase_date + timedelta(days=(warranty.warranty_period * 30))
        expiry_date = warranty.purchase_date + relativedelta(months=warranty.warranty_period) - relativedelta(days=1)
        reminder_date = expiry_date - timedelta(days=5)
        if warranty.get_status() == 'renew':
            if expiry_date >= today:
                warranties_in_range.append((warranty, expiry_date,reminder_date))

    # Sort warranties by expiry date
    # warranties_in_range.sort(key=lambda x: x.purchase_date + timedelta(days=(x.warranty_period * 30)))
    warranties_in_range.sort(key=lambda x: x[1])
    
    return warranties_in_range



@require_POST
def delete_warranty(request, id):
    try:
        warranty = get_object_or_404(Warranty, id = id)
        warranty.delete()
        messages.info(request, "Warranty Deleted Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)        
        # return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'fail', 'error': str(e)}, status=400)
    
    return render(request, "Dealer/warranty/warranty.html")
    
def warranty(request):
    user = request.user
    car_brand = CarBrands.objects.filter(is_active=True)
    warranty_list = Warranty.objects.filter(created_by=user).order_by('-purchase_date')
    warranty_count = warranty_list.count()
    warranty_id = request.POST.get('warranty_id')

    warranties_with_expiry = sorted_warranty_users(request)
    expired_warranties, live_warranty = expired_warranty(request)
    live_count = len(live_warranty)
    expired_count = len(expired_warranties)
    renewal_count = len(warranties_with_expiry)
    
    for warranty in warranty_list:
        warranty.status = warranty.get_status()
        warranty.save()
    
    if warranty_id:
        # Edit existing warranty
        warranty = get_object_or_404(Warranty, pk=warranty_id)
        warranty_period = request.POST.get('warranty_period')
        purchase_date = request.POST.get('purchase_date')
        warranty_type = request.POST.get('warrantyType')
        if request.POST.get('rsa') == "Yes":
            roadside_assistant_date = request.POST.get('dateField')
        else:
            roadside_assistant_date = None

        warranty_dict = {'warranty_period': warranty_period,'purchase_date': purchase_date, 'roadside_assistant_date': roadside_assistant_date,'warranty_type':warranty_type,'is_active': True}        
        for field, value in warranty_dict.items():
            setattr(warranty, field, value)
        warranty.save()    
    

    if request.method == 'POST' and not warranty_id:
        user_phone_no = request.POST.get('user_phone_no')
        user_name = request.POST.get('user_fname')+ " " +request.POST.get('user_mname') + " " +request.POST.get('user_lname')
        vehicle_registration_no = request.POST.get('vehicle_registration_no')
        warranty_period = request.POST.get('warranty_period')
        purchase_date = request.POST.get('purchase_date')
        if request.POST.get('rsa') == "Yes":
            roadside_assistant_date = request.POST.get('dateField')
        else:
            roadside_assistant_date = None
        warranty_type = request.POST.get('warrantyType')
        car_model = request.POST.get('car_model')
        next_renewal_date = request.POST.get('next_renewal_date') + "-01"
        car_model_instance = CarModel.objects.get(id=car_model)
        created_by = user
        warranty_dict = {'user_phone_no': user_phone_no, 'user_name': user_name, 'vehicle_registration_no': vehicle_registration_no, 'warranty_period': warranty_period,
        'purchase_date': purchase_date, 'roadside_assistant_date': roadside_assistant_date,'warranty_type':warranty_type, 'car_model': car_model_instance, 'next_renewal_date': next_renewal_date,
            'created_by': created_by,'is_active': True}
        # print('warranty_dict = ', warranty_dict)

        Warranty.objects.create(**warranty_dict)
        messages.info(request, "Warranty Form Submit Successfully !!")
        referrer = request.META.get('HTTP_REFERER')
        return HttpResponseRedirect(referrer)



    return render(request, "Dealer/warranty/warranty.html", {'car_brand': car_brand, 'warranty_list': warranty_list,'warranty_count':warranty_count,'renewal_count':renewal_count,'expired_count':expired_count,
                            'warranties_with_expiry':warranties_with_expiry,'expired_warranties': expired_warranties, 'current_warranty':live_warranty, 'live_count':live_count})



def generate_visiting_card(car_detail):
    # Create an image with white background
    img = Image.new('RGB', (800, 400), color='white')
    d = ImageDraw.Draw(img)

    # Load a font
    font_path = os.path.join(settings.BASE_DIR, 'path/to/your/font.ttf')
    font = ImageFont.truetype(font_path, 30)

    # Draw text on the image
    d.text((10, 10), f"Car Name: {car_detail.variant.model.brand.name} {car_detail.variant.model.name} {car_detail.variant.name}", font=font, fill=(0, 0, 0))
    d.text((10, 50), f"Price: {car_detail.demand_price}", font=font, fill=(0, 0, 0))
    d.text((10, 90), f"Dealer: {car_detail.dealer.name}", font=font, fill=(0, 0, 0))
    d.text((10, 130), f"Phone: {car_detail.dealer.phone}", font=font, fill=(0, 0, 0))
    d.text((10, 170), f"Registration No: {car_detail.registration_no}", font=font, fill=(0, 0, 0))
    d.text((10, 210), f"Date: {car_detail.created_at.date()}", font=font, fill=(0, 0, 0))
    d.text((10, 250), f"Status: {car_detail.status.capitalize()}", font=font, fill=(0, 0, 0))
    d.text((10, 290), f"Link: https://www.dreamcarbazaar.com/car_detail/{car_detail.id}", font=font, fill=(0, 0, 0))

    # Save the image in memory
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)

    return buffer

def generate_card_view(request, car_id):
    car_detail = get_object_or_404(CarDetail, id=car_id)
    image_buffer = generate_visiting_card(car_detail)
    return HttpResponse(image_buffer, content_type='image/png')
